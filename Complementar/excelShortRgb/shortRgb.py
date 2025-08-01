from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color
import re

# Caminho do arquivo Excel
arquivo_excel = "paintCD.xlsx"

# Carregar planilha
wb = load_workbook(arquivo_excel, data_only=True)
ws = wb.active

# Função para converter cor ARGB para RGB normal
def argb_to_rgb(color_str):
    if color_str is None:
        return (224, 224, 224)  # branco padrão
    hex_str = color_str.rgb  # Ex: 'FF00FF00' = ARGB
    if hex_str and re.fullmatch(r'[0-9A-Fa-f]{8}', hex_str):
        r = int(hex_str[2:4], 16)
        g = int(hex_str[4:6], 16)
        b = int(hex_str[6:8], 16)
        return (r, g, b)
    else:
        return (224, 224, 224)  # fallback

if __name__ == "__main__":
    rgb = []
    # Iterar 120 linhas × 160 colunas
    for row in range(1, 121):
        for col in range(1, 161):
            cell = ws.cell(row=row, column=col)
            fill = cell.fill
            color = 224, 224, 224  # branco padrão

            if fill and fill.fgColor and fill.fgColor.type == "rgb":
                color = argb_to_rgb(fill.fgColor)

            # Redução de 0–255 para 0–7
            red   = color[0] // 32
            green = color[1] // 32
            blue  = color[2] // 32
            rgb.append((red, green, blue))

    with open('framebuffer.mif', 'w') as fd:
        fd.write("""WIDTH=9;
DEPTH=19200;

ADDRESS_RADIX=UNS;
DATA_RADIX=BIN;

CONTENT BEGIN
""")
        for it, pixel in enumerate(rgb):
            fd.write(f"{it}: {format(pixel[0], '03b')}{format(pixel[1], '03b')}{format(pixel[2], '03b')};\n")
            it = it + 1
        fd.write("END")

    print("EOP")
